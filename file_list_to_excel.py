import os
from openpyxl import Workbook, load_workbook

# === Configuration ===
excel_filename = "file_list.xlsx"
duplicate_output = "file_list_duplicates.xlsx"

# Extensions: lowercase, no dot
blacklist = ["py", "sh"]
whitelist = ["mp4"]  # Leave empty list to disable

# === Setup ===
directory = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(directory, excel_filename)
dup_path = os.path.join(directory, duplicate_output)

def is_valid_file(filename):
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if whitelist:
        return ext in whitelist and ext not in blacklist
    return ext not in blacklist

# Get filtered files from current directory
all_files = [f for f in os.listdir(directory)
             if os.path.isfile(os.path.join(directory, f)) and f != excel_filename]

filtered_files = [f for f in all_files if is_valid_file(f)]

# === Load main workbook ===
if os.path.exists(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "File List"
    ws.append(["Filename"])  # Default header

# === Read data ===
rows = list(ws.iter_rows(values_only=True))
header = rows[0] if rows else ["Filename"]
data_rows = rows[1:]

# === Preserve empty rows ===
non_empty_rows = []
empty_rows = []

for row in data_rows:
    if not any(row):  # All cells empty
        empty_rows.append(row)
    else:
        non_empty_rows.append(row)

# === Detect duplicates in existing data ===
filename_map = {}
duplicates = {}

for row in non_empty_rows:
    filename = row[0]
    if filename in filename_map:
        if filename not in duplicates:
            duplicates[filename] = [filename_map[filename]]  # Add the first seen
        duplicates[filename].append(row)  # Add the new one
    else:
        filename_map[filename] = row

# === Add new files if not present ===
for f in filtered_files:
    if f not in filename_map and f not in duplicates:
        filename_map[f] = (f,) + ("",) * (len(header) - 1)

# === Sort remaining (non-duplicate) rows by filename ===
sorted_rows = sorted(filename_map.values(), key=lambda row: row[0].lower())

# === Rewrite main Excel file ===
ws.delete_rows(2, ws.max_row)  # Keep header

for row in sorted_rows:
    ws.append(row)

for row in empty_rows:
    ws.append(row)

wb.save(excel_path)
print(f"✅ Main Excel updated. {len(sorted_rows)} files listed, {len(empty_rows)} empty rows preserved.")
print(f"⚠️ Found {len(duplicates)} duplicate filename conflict(s). Writing them to a new file...")

# === Write duplicates to a new workbook ===
dup_wb = Workbook()
dup_ws = dup_wb.active
dup_ws.title = "Duplicates"
dup_ws.append(header)

for filename, conflict_rows in duplicates.items():
    for row in conflict_rows:
        dup_ws.append(row)
    dup_ws.append(())  # Empty row between groups

dup_wb.save(dup_path)
print(f"📄 Duplicates written to: {duplicate_output}. Please review them manually.")
